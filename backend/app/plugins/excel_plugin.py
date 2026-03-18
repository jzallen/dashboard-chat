"""Excel file format plugin."""

import io
from typing import ClassVar

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .protocol import PluginChoice, PluginValidationError, ProcessingResult


class ExcelPlugin:
    """Plugin for Excel (.xlsx, .xls) file processing."""

    name: ClassVar[str] = "excel"
    extensions: ClassVar[list[str]] = [".xlsx", ".xls"]
    label: ClassVar[str] = "Excel"
    dbt_macros: ClassVar[dict[str, str] | None] = None

    def validate(self, file_content: bytes, filename: str) -> None:
        if not file_content:
            raise PluginValidationError("File is empty")
        try:
            load_workbook(io.BytesIO(file_content), read_only=True)
        except (InvalidFileException, Exception) as e:
            raise PluginValidationError("Invalid Excel file") from e

    def detect_choices(self, file_content: bytes, filename: str) -> list[PluginChoice] | None:
        wb = load_workbook(io.BytesIO(file_content), read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        if len(sheet_names) > 1:
            return [
                PluginChoice(
                    key="sheet_name",
                    label="Select a sheet to import",
                    options=sheet_names,
                )
            ]
        return None

    def process(self, file_content: bytes, filename: str, choices: dict[str, str] | None = None) -> ProcessingResult:
        selected = None
        if choices and "sheet_name" in choices:
            selected = choices["sheet_name"]

        if selected is None:
            # Use the first sheet
            selected = 0

        try:
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=selected, engine="openpyxl")
        except ValueError as e:
            raise PluginValidationError(f"Invalid sheet name: {e}") from e

        df.columns = df.columns.str.strip()
        str_cols = df.select_dtypes(include="object").columns
        df[str_cols] = df[str_cols].apply(lambda col: col.str.strip())
        return ProcessingResult(df=df)
