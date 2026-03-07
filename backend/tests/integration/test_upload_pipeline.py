"""Integration tests for the upload → dataset pipeline across file formats.

Tests the full flow: upload_file → create_dataset_from_upload for each plugin.
Uses seeded_db and S3 stubbers (no external services required).
"""

import io
import json
import zipfile
from functools import partial
from io import BytesIO

import boto3
import pytest
from botocore.stub import Stubber
from openpyxl import Workbook
from returns.result import Success
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth.context import set_auth_user
from app.auth.types import AuthUser
from app.models import Upload
from app.models.dataset import Dataset
from app.plugins import PluginRegistry, create_plugin_registry
from app.repositories import set_session
from app.repositories.lake import MinIOLakeRepository
from app.repositories.metadata import ProjectRecord
from app.use_cases.dataset import create_dataset_from_upload
from app.use_cases.project._dbt import generate_dbt_project_zip
from app.use_cases.upload import upload_file
from tests.uuidv7_fixtures import ORG_1, PROJECT_1, USER_1

TEST_USER = AuthUser(id=USER_1, email="test@example.com", org_id=ORG_1, name="Test User")


@pytest.fixture(autouse=True)
def auth_user():
    set_auth_user(TEST_USER)


@pytest.fixture
async def seeded_db(db_session: AsyncSession):
    """Seed with a project."""
    db_session.add(ProjectRecord(id=PROJECT_1, name="Test Project", org_id=ORG_1))
    await db_session.commit()
    return db_session


@pytest.fixture
def plugin_registry():
    return create_plugin_registry()


def _make_s3_stubber_for_upload(file_content: bytes, storage_path: str, num_partitions: int = 1) -> Stubber:
    """Create an S3 stubber for the full upload → dataset flow.

    Stubs: put_object (raw upload), get_object (read back), put_object * N (parquet writes).
    """
    stubber = Stubber(boto3.client("s3"))
    # upload_file writes raw file
    stubber.add_response("put_object", {})
    # create_dataset_from_upload reads raw file back
    stubber.add_response(
        "get_object",
        {"Body": io.BytesIO(file_content)},
        {"Bucket": "dashboard-chat.datalake", "Key": storage_path},
    )
    # create_dataset_from_upload writes parquet partition(s)
    for _ in range(num_partitions):
        stubber.add_response("put_object", {})
    return stubber


def _make_excel_single_sheet() -> bytes:
    """Create a single-sheet Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["name", "age"])
    ws.append(["Alice", 30])
    ws.append(["Bob", 25])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_excel_multi_sheet() -> bytes:
    """Create a multi-sheet Excel file."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Patients"
    ws1.append(["name", "age"])
    ws1.append(["Alice", 30])
    ws2 = wb.create_sheet("Visits")
    ws2.append(["date", "type"])
    ws2.append(["2024-01-01", "checkup"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_hl7_content() -> bytes:
    """Create a simple HL7v2 message."""
    return (
        b"MSH|^~\\&|SndApp|SndFac|RcvApp|RcvFac|20240101120000||ADT^A01|MSG001|P|2.3\r"
        b"PID|||12345^^^MRN||Doe^John||19800101|M\r"
        b"PV1||I|ICU^101^A\r"
    )


def _make_fhir_bundle_multi_type() -> bytes:
    """Create a FHIR bundle with Patient and Observation resources."""
    bundle = {
        "resourceType": "Bundle",
        "entry": [
            {"resource": {"resourceType": "Patient", "id": "p1", "gender": "male", "birthDate": "1980-01-01"}},
            {
                "resource": {
                    "resourceType": "Observation",
                    "id": "o1",
                    "status": "final",
                    "code": {"text": "Blood Pressure"},
                }
            },
            {"resource": {"resourceType": "Patient", "id": "p2", "gender": "female", "birthDate": "1990-05-15"}},
        ],
    }
    return json.dumps(bundle).encode()


def _make_fhir_bundle_single_type() -> bytes:
    """Create a FHIR bundle with only Patient resources."""
    bundle = {
        "resourceType": "Bundle",
        "entry": [
            {"resource": {"resourceType": "Patient", "id": "1", "gender": "male", "birthDate": "1980-01-01"}},
            {"resource": {"resourceType": "Patient", "id": "2", "gender": "female", "birthDate": "1990-05-15"}},
        ],
    }
    return json.dumps(bundle).encode()


def _make_fhir_ndjson_multi_type() -> bytes:
    """Create NDJSON with multiple FHIR resource types."""
    resources = [
        {"resourceType": "Patient", "id": "1", "gender": "male"},
        {"resourceType": "Observation", "id": "2", "status": "final", "code": {"text": "BP"}},
        {"resourceType": "Patient", "id": "3", "gender": "female"},
    ]
    return "\n".join(json.dumps(r) for r in resources).encode()


class TestCsvPipeline:
    """12.1: CSV upload → dataset created (regression)."""

    async def test_csv_upload_creates_dataset(self, seeded_db: AsyncSession, plugin_registry: PluginRegistry):
        set_session(seeded_db)
        csv_content = b"name,age,active\nAlice,30,true\nBob,25,false"

        # Step 1: Upload
        write_stubber = Stubber(boto3.client("s3"))
        write_stubber.add_response("put_object", {})
        with write_stubber:
            upload_result = await upload_file(
                file_content=csv_content,
                file_name="people.csv",
                project_id=PROJECT_1,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=write_stubber.client)},
            )

        assert isinstance(upload_result, Success)
        upload = upload_result.unwrap()
        assert isinstance(upload, Upload)
        assert upload.status == "pending"
        assert len(upload.preview_rows) == 2

        # Step 2: Create dataset
        upload_id = upload.id
        raw_path = upload.raw_storage_path
        read_write_stubber = Stubber(boto3.client("s3"))
        read_write_stubber.add_response(
            "get_object",
            {"Body": io.BytesIO(csv_content)},
            {"Bucket": "dashboard-chat.datalake", "Key": raw_path},
        )
        read_write_stubber.add_response("put_object", {})  # parquet write

        with read_write_stubber:
            dataset_result = await create_dataset_from_upload(
                upload_id=upload_id,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=read_write_stubber.client)},
            )

        assert isinstance(dataset_result, Success)
        dataset = dataset_result.unwrap()
        assert isinstance(dataset, Dataset)
        assert set(dataset.schema_config["fields"].keys()) == {"name", "age", "active"}
        assert len(dataset.preview_rows) == 2


class TestExcelSingleSheetPipeline:
    """12.2: Excel single-sheet upload → dataset created."""

    async def test_excel_single_sheet_creates_dataset(self, seeded_db: AsyncSession, plugin_registry: PluginRegistry):
        set_session(seeded_db)
        excel_content = _make_excel_single_sheet()

        # Step 1: Upload
        write_stubber = Stubber(boto3.client("s3"))
        write_stubber.add_response("put_object", {})
        with write_stubber:
            upload_result = await upload_file(
                file_content=excel_content,
                file_name="data.xlsx",
                project_id=PROJECT_1,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=write_stubber.client)},
            )

        assert isinstance(upload_result, Success)
        upload = upload_result.unwrap()
        assert upload.status == "pending"  # single sheet, no choices needed
        assert len(upload.preview_rows) == 2

        # Step 2: Create dataset
        upload_id = upload.id
        raw_path = upload.raw_storage_path
        read_write_stubber = Stubber(boto3.client("s3"))
        read_write_stubber.add_response(
            "get_object",
            {"Body": io.BytesIO(excel_content)},
            {"Bucket": "dashboard-chat.datalake", "Key": raw_path},
        )
        read_write_stubber.add_response("put_object", {})

        with read_write_stubber:
            dataset_result = await create_dataset_from_upload(
                upload_id=upload_id,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=read_write_stubber.client)},
            )

        assert isinstance(dataset_result, Success)
        dataset = dataset_result.unwrap()
        assert set(dataset.schema_config["fields"].keys()) == {"name", "age"}


class TestExcelMultiSheetPipeline:
    """12.3: Excel multi-sheet → choices → select → dataset created."""

    async def test_excel_multi_sheet_with_choices_creates_dataset(
        self, seeded_db: AsyncSession, plugin_registry: PluginRegistry
    ):
        set_session(seeded_db)
        excel_content = _make_excel_multi_sheet()

        # Step 1: Upload → should return awaiting_input with choices
        write_stubber = Stubber(boto3.client("s3"))
        write_stubber.add_response("put_object", {})
        with write_stubber:
            upload_result = await upload_file(
                file_content=excel_content,
                file_name="multi.xlsx",
                project_id=PROJECT_1,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=write_stubber.client)},
            )

        assert isinstance(upload_result, Success)
        upload = upload_result.unwrap()
        assert upload.status == "awaiting_input"
        assert upload.choices is not None
        assert len(upload.choices) > 0
        # Verify sheet names are in choices
        choice_options = upload.choices[0].get("options", [])
        assert "Patients" in choice_options
        assert "Visits" in choice_options

        # Step 2: Process with user's choice (select "Patients" sheet)
        upload_id = upload.id
        raw_path = upload.raw_storage_path
        read_write_stubber = Stubber(boto3.client("s3"))
        read_write_stubber.add_response(
            "get_object",
            {"Body": io.BytesIO(excel_content)},
            {"Bucket": "dashboard-chat.datalake", "Key": raw_path},
        )
        read_write_stubber.add_response("put_object", {})

        with read_write_stubber:
            dataset_result = await create_dataset_from_upload(
                upload_id=upload_id,
                plugin_registry=plugin_registry,
                choices={"sheet_name": "Patients"},
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=read_write_stubber.client)},
            )

        assert isinstance(dataset_result, Success)
        dataset = dataset_result.unwrap()
        assert set(dataset.schema_config["fields"].keys()) == {"name", "age"}


class TestHl7v2Pipeline:
    """12.4: HL7 file upload → requires Mirth Connect, validates MSH."""

    async def test_hl7_upload_validates_mirth_connect_required(
        self, seeded_db: AsyncSession, plugin_registry: PluginRegistry
    ):
        """HL7v2 upload should fail validation when Mirth Connect URL is not configured."""
        set_session(seeded_db)
        hl7_content = _make_hl7_content()

        # Step 1: Upload — should fail because Mirth Connect URL is not set
        write_stubber = Stubber(boto3.client("s3"))
        write_stubber.add_response("put_object", {})
        with write_stubber:
            upload_result = await upload_file(
                file_content=hl7_content,
                file_name="messages.hl7",
                project_id=PROJECT_1,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=write_stubber.client)},
            )

        # Expect failure because Mirth Connect is not configured
        from returns.result import Failure as F

        assert isinstance(upload_result, F)


class TestHl7v2EndToEndPipeline:
    """W2: HL7v2 upload → mocked Mirth conversion → FHIR → multi-dataset output."""

    async def test_hl7_upload_with_mocked_mirth_creates_multi_datasets(
        self, seeded_db: AsyncSession
    ):
        """Full HL7v2 E2E: upload → Mirth converts to FHIR → FHIR plugin splits → multiple datasets."""
        from unittest.mock import MagicMock, patch

        set_session(seeded_db)
        hl7_content = _make_hl7_content()

        # Mock FHIR bundle that Mirth would return
        mock_fhir_bundle = {
            "resourceType": "Bundle",
            "type": "transaction",
            "entry": [
                {
                    "resource": {
                        "resourceType": "Patient",
                        "id": "p1",
                        "gender": "male",
                        "birthDate": "1980-01-01",
                    }
                },
                {
                    "resource": {
                        "resourceType": "Encounter",
                        "id": "e1",
                        "status": "finished",
                        "class": [{"text": "inpatient"}],
                    }
                },
            ],
        }

        mock_settings = MagicMock()
        mock_settings.mirth_connect_url = "http://mirth:8443"
        mock_settings.mirth_connect_api_key = "test-key"
        mock_settings.mirth_connect_timeout = 60

        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.json.return_value = mock_fhir_bundle

        plugin_registry = create_plugin_registry()

        with patch("app.plugins.hl7v2_plugin.get_settings", return_value=mock_settings), \
             patch("app.plugins.mirth_client.httpx.post", return_value=mock_response):

            # Step 1: Upload raw HL7v2 file
            write_stubber = Stubber(boto3.client("s3"))
            write_stubber.add_response("put_object", {})
            with write_stubber:
                upload_result = await upload_file(
                    file_content=hl7_content,
                    file_name="messages.hl7",
                    project_id=PROJECT_1,
                    plugin_registry=plugin_registry,
                    repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=write_stubber.client)},
                )

            assert isinstance(upload_result, Success)
            upload = upload_result.unwrap()
            assert upload.status == "pending"

            # Step 2: Create datasets from uploaded HL7v2
            raw_path = upload.raw_storage_path
            read_write_stubber = Stubber(boto3.client("s3"))
            read_write_stubber.add_response(
                "get_object",
                {"Body": io.BytesIO(hl7_content)},
                {"Bucket": "dashboard-chat.datalake", "Key": raw_path},
            )
            # put_object for converted FHIR artifact
            read_write_stubber.add_response("put_object", {})
            # 2 datasets × 1 parquet partition each
            read_write_stubber.add_response("put_object", {})
            read_write_stubber.add_response("put_object", {})

            with read_write_stubber:
                dataset_result = await create_dataset_from_upload(
                    upload_id=upload.id,
                    plugin_registry=plugin_registry,
                    repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=read_write_stubber.client)},
                )

        assert isinstance(dataset_result, Success)
        datasets = dataset_result.unwrap()
        assert isinstance(datasets, list)
        assert len(datasets) == 2

        names = {d.name for d in datasets}
        assert "Patient" in names
        assert "Encounter" in names

        for d in datasets:
            assert d.project_id == PROJECT_1
            assert len(d.preview_rows) > 0


class TestFhirBundlePipeline:
    """8.1: FHIR R4 Bundle upload → multiple datasets created with correct resource types."""

    async def test_fhir_bundle_creates_multiple_datasets(
        self, seeded_db: AsyncSession, plugin_registry: PluginRegistry
    ):
        set_session(seeded_db)
        fhir_content = _make_fhir_ndjson_multi_type()  # Patient + Observation

        # Step 1: Upload → no choices (FHIR auto-splits by resource type)
        write_stubber = Stubber(boto3.client("s3"))
        write_stubber.add_response("put_object", {})
        with write_stubber:
            upload_result = await upload_file(
                file_content=fhir_content,
                file_name="bundle.ndjson",
                project_id=PROJECT_1,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=write_stubber.client)},
            )

        assert isinstance(upload_result, Success)
        upload = upload_result.unwrap()
        assert upload.status == "pending"  # No choices needed
        assert upload.choices is None
        assert len(upload.preview_rows) > 0  # Preview from first resource type

        # Step 2: Create datasets — should produce one per resource type
        upload_id = upload.id
        raw_path = upload.raw_storage_path
        read_write_stubber = Stubber(boto3.client("s3"))
        read_write_stubber.add_response(
            "get_object",
            {"Body": io.BytesIO(fhir_content)},
            {"Bucket": "dashboard-chat.datalake", "Key": raw_path},
        )
        # 2 datasets × 1 parquet partition each
        read_write_stubber.add_response("put_object", {})
        read_write_stubber.add_response("put_object", {})

        with read_write_stubber:
            dataset_result = await create_dataset_from_upload(
                upload_id=upload_id,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=read_write_stubber.client)},
            )

        assert isinstance(dataset_result, Success)
        datasets = dataset_result.unwrap()
        assert isinstance(datasets, list)
        assert len(datasets) == 2

        names = {d.name for d in datasets}
        assert "Observation" in names
        assert "Patient" in names

        for d in datasets:
            assert d.project_id == PROJECT_1
            assert d.format_context is not None
            assert len(d.preview_rows) > 0

    async def test_fhir_single_type_bundle_creates_single_item_list(
        self, seeded_db: AsyncSession, plugin_registry: PluginRegistry
    ):
        """Single-type FHIR bundle still returns list (MultiProcessingResult always)."""
        set_session(seeded_db)
        fhir_content = _make_fhir_bundle_single_type()  # Only Patient

        # Step 1: Upload
        write_stubber = Stubber(boto3.client("s3"))
        write_stubber.add_response("put_object", {})
        with write_stubber:
            upload_result = await upload_file(
                file_content=fhir_content,
                file_name="patients.fhir.json",
                project_id=PROJECT_1,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=write_stubber.client)},
            )

        assert isinstance(upload_result, Success)
        upload = upload_result.unwrap()
        assert upload.status == "pending"

        # Step 2: Create dataset
        upload_id = upload.id
        raw_path = upload.raw_storage_path
        read_write_stubber = Stubber(boto3.client("s3"))
        read_write_stubber.add_response(
            "get_object",
            {"Body": io.BytesIO(fhir_content)},
            {"Bucket": "dashboard-chat.datalake", "Key": raw_path},
        )
        read_write_stubber.add_response("put_object", {})

        with read_write_stubber:
            dataset_result = await create_dataset_from_upload(
                upload_id=upload_id,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=read_write_stubber.client)},
            )

        assert isinstance(dataset_result, Success)
        datasets = dataset_result.unwrap()
        assert isinstance(datasets, list)
        assert len(datasets) == 1
        assert datasets[0].name == "Patient"


class TestCsvUnchangedBehavior:
    """8.3: CSV upload → single dataset unchanged behavior."""

    async def test_csv_upload_returns_single_dataset_not_list(
        self, seeded_db: AsyncSession, plugin_registry: PluginRegistry
    ):
        """CSV plugin returns ProcessingResult, so create_dataset_from_upload returns a single Dataset."""
        set_session(seeded_db)
        csv_content = b"x,y\n1,2\n3,4"

        # Upload
        write_stubber = Stubber(boto3.client("s3"))
        write_stubber.add_response("put_object", {})
        with write_stubber:
            upload_result = await upload_file(
                file_content=csv_content,
                file_name="simple.csv",
                project_id=PROJECT_1,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=write_stubber.client)},
            )

        assert isinstance(upload_result, Success)
        upload = upload_result.unwrap()

        # Create dataset
        raw_path = upload.raw_storage_path
        read_write_stubber = Stubber(boto3.client("s3"))
        read_write_stubber.add_response(
            "get_object",
            {"Body": io.BytesIO(csv_content)},
            {"Bucket": "dashboard-chat.datalake", "Key": raw_path},
        )
        read_write_stubber.add_response("put_object", {})

        with read_write_stubber:
            dataset_result = await create_dataset_from_upload(
                upload_id=upload.id,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=read_write_stubber.client)},
            )

        assert isinstance(dataset_result, Success)
        dataset = dataset_result.unwrap()
        # CSV returns single Dataset, NOT a list
        assert isinstance(dataset, Dataset)
        assert not isinstance(dataset, list)


class TestMultiDatasetResponseShape:
    """8.4: Upload API response includes dataset_ids for multi-dataset upload."""

    async def test_multi_dataset_upload_response_has_dataset_ids(
        self, seeded_db: AsyncSession, plugin_registry: PluginRegistry
    ):
        """Multi-dataset result serializes as list of dataset dicts."""
        set_session(seeded_db)
        fhir_content = _make_fhir_ndjson_multi_type()  # Patient + Observation

        # Upload
        write_stubber = Stubber(boto3.client("s3"))
        write_stubber.add_response("put_object", {})
        with write_stubber:
            upload_result = await upload_file(
                file_content=fhir_content,
                file_name="multi.ndjson",
                project_id=PROJECT_1,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=write_stubber.client)},
            )

        assert isinstance(upload_result, Success)
        upload = upload_result.unwrap()

        # Verify upload serialization includes new fields
        serialized = upload.serialize()
        assert "dataset_ids" in serialized
        assert "converted_storage_path" in serialized
        assert "dataset_id" in serialized  # backward compat

        # Create datasets
        raw_path = upload.raw_storage_path
        read_write_stubber = Stubber(boto3.client("s3"))
        read_write_stubber.add_response(
            "get_object",
            {"Body": io.BytesIO(fhir_content)},
            {"Bucket": "dashboard-chat.datalake", "Key": raw_path},
        )
        read_write_stubber.add_response("put_object", {})
        read_write_stubber.add_response("put_object", {})

        with read_write_stubber:
            dataset_result = await create_dataset_from_upload(
                upload_id=upload.id,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=read_write_stubber.client)},
            )

        assert isinstance(dataset_result, Success)
        datasets = dataset_result.unwrap()
        assert isinstance(datasets, list)

        # Each dataset serializes correctly
        for d in datasets:
            s = d.serialize()
            assert "id" in s
            assert "project_id" in s
            assert "name" in s
            assert "schema_config" in s


class TestFhirMultiDatasetResponseShape:
    """6.3: Upload FHIR bundle → verify multi-dataset response shape.

    Validates:
    - create_dataset_from_upload returns list[Dataset] with correct fields
    - Upload outbox record is updated with dataset_ids (all) and dataset_id (first, backward compat)
    - Controller _serialize produces correct HTTP response structure
    """

    async def test_fhir_bundle_updates_outbox_with_dataset_ids(
        self, seeded_db: AsyncSession, plugin_registry: PluginRegistry
    ):
        """After multi-dataset creation, the outbox record should contain dataset_ids and dataset_id."""
        set_session(seeded_db)
        fhir_content = _make_fhir_bundle_multi_type()

        # Step 1: Upload (.ndjson triggers FHIR plugin; _parse_fhir_content handles JSON bundles too)
        write_stubber = Stubber(boto3.client("s3"))
        write_stubber.add_response("put_object", {})
        with write_stubber:
            upload_result = await upload_file(
                file_content=fhir_content,
                file_name="bundle.ndjson",
                project_id=PROJECT_1,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=write_stubber.client)},
            )

        assert isinstance(upload_result, Success)
        upload = upload_result.unwrap()
        upload_id = upload.id
        raw_path = upload.raw_storage_path

        # Step 2: Create datasets (Patient + Observation = 2 parquet writes)
        read_write_stubber = Stubber(boto3.client("s3"))
        read_write_stubber.add_response(
            "get_object",
            {"Body": io.BytesIO(fhir_content)},
            {"Bucket": "dashboard-chat.datalake", "Key": raw_path},
        )
        read_write_stubber.add_response("put_object", {})
        read_write_stubber.add_response("put_object", {})

        with read_write_stubber:
            dataset_result = await create_dataset_from_upload(
                upload_id=upload_id,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=read_write_stubber.client)},
            )

        assert isinstance(dataset_result, Success)
        datasets = dataset_result.unwrap()
        assert isinstance(datasets, list)
        dataset_ids = [d.id for d in datasets]
        assert len(set(dataset_ids)) == 2, "Each dataset should have a distinct ID"

        # Verify outbox record was updated with dataset_ids and dataset_id
        from app.repositories.outbox import OutboxRecord

        record = await seeded_db.get(OutboxRecord, upload_id)
        assert record is not None
        assert record.payload.get("dataset_ids") == dataset_ids
        assert record.payload.get("dataset_id") == dataset_ids[0], (
            "dataset_id should be first ID for backward compat"
        )

    async def test_fhir_multi_dataset_controller_serialize_shape(
        self, seeded_db: AsyncSession, plugin_registry: PluginRegistry
    ):
        """The controller _serialize path should wrap list[Dataset] as list of dicts in response."""
        set_session(seeded_db)
        fhir_content = _make_fhir_bundle_multi_type()

        # Upload
        write_stubber = Stubber(boto3.client("s3"))
        write_stubber.add_response("put_object", {})
        with write_stubber:
            upload_result = await upload_file(
                file_content=fhir_content,
                file_name="bundle.ndjson",
                project_id=PROJECT_1,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=write_stubber.client)},
            )
        upload = upload_result.unwrap()

        # Create datasets
        read_write_stubber = Stubber(boto3.client("s3"))
        read_write_stubber.add_response(
            "get_object",
            {"Body": io.BytesIO(fhir_content)},
            {"Bucket": "dashboard-chat.datalake", "Key": upload.raw_storage_path},
        )
        read_write_stubber.add_response("put_object", {})
        read_write_stubber.add_response("put_object", {})

        with read_write_stubber:
            dataset_result = await create_dataset_from_upload(
                upload_id=upload.id,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=read_write_stubber.client)},
            )

        datasets = dataset_result.unwrap()

        # Simulate controller serialization (same path as HTTPController.post_dataset)
        from app.controllers.http_controller import _serialize
        from app.controllers.response_wrapper import wrap_success

        response_body = wrap_success(_serialize(datasets))

        # Verify top-level response shape
        assert response_body["success"] is True
        assert isinstance(response_body["data"], list)
        assert len(response_body["data"]) == 2

        # Verify each dataset dict has all expected fields
        names = set()
        for item in response_body["data"]:
            assert isinstance(item, dict)
            assert "id" in item
            assert "project_id" in item
            assert item["project_id"] == PROJECT_1
            assert "name" in item
            assert "schema_config" in item
            assert "format_context" in item
            assert "staging_sql" in item
            assert "preview_rows" in item
            names.add(item["name"])

        assert names == {"Observation", "Patient"}


class TestDbtExportWithMixedFormats:
    """12.6: dbt export with mixed-format datasets → valid zip with plugin macros."""

    def test_export_with_plugin_macros_produces_valid_zip(self, plugin_registry: PluginRegistry):
        from app.models.project import Project

        # Create project with datasets that have format_context
        ds_csv = Dataset(
            id="ds-csv",
            project_id="proj-1",
            name="CSV Data",
            schema_config={"fields": {"name": {"type": "text"}}},
            transforms=[],
        )
        ds_hl7 = Dataset(
            id="ds-hl7",
            project_id="proj-1",
            name="HL7 Messages",
            schema_config={"fields": {"msh_3": {"type": "text"}, "pid_5": {"type": "text"}}},
            transforms=[],
            format_context="HL7v2 format context",
        )
        project = Project(id="proj-1", name="Mixed Format Project", datasets=[ds_csv, ds_hl7])

        zip_bytes = generate_dbt_project_zip(project, "mixed_format_project", plugin_registry=plugin_registry)
        zf = zipfile.ZipFile(BytesIO(zip_bytes))
        names = set(zf.namelist())

        # Core files present
        assert "dbt_project.yml" in names
        assert "models/staging/stg_csv_data.sql" in names
        assert "models/staging/stg_hl7_messages.sql" in names

        # Plugin macros present (HL7v2 plugin has dbt_macros)
        plugin_macro_files = [n for n in names if n.startswith("macros/plugin_")]
        assert len(plugin_macro_files) > 0, "Expected plugin macro files in zip"

        # Verify zip is valid
        for name in zf.namelist():
            zf.read(name).decode("utf-8")  # All files should be valid UTF-8
