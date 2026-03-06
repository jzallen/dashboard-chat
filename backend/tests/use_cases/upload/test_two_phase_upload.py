"""Tests for the two-phase upload flow (detect choices → process with choices).

Tests the upload_file use case returning awaiting_input status when
a plugin detects choices, and the subsequent create_dataset_from_upload
with user-provided choices.
"""

import io
from functools import partial
from io import BytesIO

import boto3
import pytest
from botocore.stub import Stubber
from openpyxl import Workbook
from returns.result import Success
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.dataset import Dataset
from app.plugins import create_plugin_registry
from app.repositories import set_session
from app.repositories.lake import MinIOLakeRepository
from app.use_cases.dataset import create_dataset_from_upload
from app.use_cases.upload import upload_file
from tests.uuidv7_fixtures import PROJECT_1


@pytest.fixture
def plugin_registry():
    return create_plugin_registry()


def _make_multi_sheet_excel() -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Orders"
    ws1.append(["order_id", "amount"])
    ws1.append([1, 100.0])
    ws2 = wb.create_sheet("Returns")
    ws2.append(["return_id", "reason"])
    ws2.append([1, "defective"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestTwoPhaseUpload:
    """Tests for the two-phase upload flow with plugin choices."""

    async def test_upload_with_choices_returns_awaiting_input(self, seeded_db: AsyncSession, plugin_registry):
        """upload_file should return awaiting_input when detect_choices returns options."""
        set_session(seeded_db)
        excel_content = _make_multi_sheet_excel()

        stubber = Stubber(boto3.client("s3"))
        stubber.add_response("put_object", {})
        with stubber:
            result = await upload_file(
                file_content=excel_content,
                file_name="report.xlsx",
                project_id=PROJECT_1,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=stubber.client)},
            )

        assert isinstance(result, Success)
        upload = result.unwrap()
        assert upload.status == "awaiting_input"
        assert upload.choices is not None
        assert len(upload.choices) == 1
        assert upload.choices[0]["key"] == "sheet_name"
        assert "Orders" in upload.choices[0]["options"]
        assert "Returns" in upload.choices[0]["options"]
        assert upload.preview_rows == []

    async def test_process_with_choices_creates_dataset(self, seeded_db: AsyncSession, plugin_registry):
        """create_dataset_from_upload with choices should process the selected sheet."""
        set_session(seeded_db)
        excel_content = _make_multi_sheet_excel()

        # Step 1: Upload
        upload_stubber = Stubber(boto3.client("s3"))
        upload_stubber.add_response("put_object", {})
        with upload_stubber:
            upload_result = await upload_file(
                file_content=excel_content,
                file_name="report.xlsx",
                project_id=PROJECT_1,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=upload_stubber.client)},
            )

        upload = upload_result.unwrap()

        # Step 2: Process with choices
        dataset_stubber = Stubber(boto3.client("s3"))
        dataset_stubber.add_response(
            "get_object",
            {"Body": io.BytesIO(excel_content)},
            {"Bucket": "dashboard-chat.datalake", "Key": upload.raw_storage_path},
        )
        dataset_stubber.add_response("put_object", {})

        with dataset_stubber:
            dataset_result = await create_dataset_from_upload(
                upload_id=upload.id,
                plugin_registry=plugin_registry,
                choices={"sheet_name": "Returns"},
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=dataset_stubber.client)},
            )

        assert isinstance(dataset_result, Success)
        dataset = dataset_result.unwrap()
        assert isinstance(dataset, Dataset)
        assert set(dataset.schema_config["fields"].keys()) == {"return_id", "reason"}

    async def test_upload_without_choices_returns_pending(self, seeded_db: AsyncSession, plugin_registry):
        """upload_file for a single-sheet Excel should return pending (not awaiting_input)."""
        set_session(seeded_db)

        wb = Workbook()
        ws = wb.active
        ws.append(["x", "y"])
        ws.append([1, 2])
        buf = BytesIO()
        wb.save(buf)
        single_sheet_content = buf.getvalue()

        stubber = Stubber(boto3.client("s3"))
        stubber.add_response("put_object", {})
        with stubber:
            result = await upload_file(
                file_content=single_sheet_content,
                file_name="single.xlsx",
                project_id=PROJECT_1,
                plugin_registry=plugin_registry,
                repositories={"lake_repository": partial(MinIOLakeRepository, s3_client=stubber.client)},
            )

        assert isinstance(result, Success)
        upload = result.unwrap()
        assert upload.status == "pending"
        assert upload.choices is None
        assert len(upload.preview_rows) == 1
